import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def extract_info(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        data = file.read().strip()  # Удаляем лишние символы в конце строки

    arm_name_match = re.search(r'Host Name:\s+(.*)', data)
    arm_name = arm_name_match.group(1) if arm_name_match else ''

    update_date_match = re.search(r'System Boot Time:\s+(.*)', data)
    update_date = update_date_match.group(1) if update_date_match else ''

    user_match = re.search(r'Executed as user:\s+(.*)', data)
    user = user_match.group(1) if user_match else ''

    ip_address_match = re.search(r'IP address\(es\)\s*[\n\r]+\s*\[.*?\]\s*:\s*([\d\.:a-fA-F]+)', data)
    ip_address = ip_address_match.group(1) if ip_address_match else ''

    os_name_match = re.search(r'OS Name:\s+(.*)', data)
    os_name = os_name_match.group(1) if os_name_match else ''

    os_version_match = re.search(r'OS Version:\s+(.*)', data)
    os_version = os_version_match.group(1) if os_version_match else ''

    os_install_date_match = re.search(r'Original Install Date:\s+(.*)', data)
    os_install_date = os_install_date_match.group(1) if os_install_date_match else ''

    cpu_match = re.search(r'CPU:\s+Name\s+([^\n]*)', data, re.DOTALL)
    cpu = cpu_match.group(1).strip() if cpu_match else ''

    system_manufacturer_match = re.search(r'System Manufacturer:\s+(.*)', data)
    system_manufacturer = system_manufacturer_match.group(1) if system_manufacturer_match else ''

    model_match = re.search(r'System Model:\s+(.*)', data)
    model = model_match.group(1) if model_match else ''

    total_memory_match = re.search(r'Total Physical Memory:\s+(\d+)', data)
    total_memory = total_memory_match.group(1) if total_memory_match else ''

    installed_ram_slots_match = re.search(r'busy slots (\d+)', data)
    installed_ram_slots = installed_ram_slots_match.group(1) if installed_ram_slots_match else ''

    free_ram_slots_match = re.search(r'Free slots (\d+)', data)
    free_ram_slots = free_ram_slots_match.group(1) if free_ram_slots_match else ''

    hdd_count = 0
    hdd_size = 0
    hdd_pattern = r'InterfaceType\s+(\S+)\s+([^\n]+)\s+(\d+)'
    for match in re.finditer(hdd_pattern, data):
        if match.group(1) != 'USB':
            hdd_count += 1
            hdd_size += int(match.group(3))

    hdd_size_gb = hdd_size / (1024 * 1024 * 1024)

    optical_drive_match = re.search(r'CD/DVD:\s+(.*)', data)
    optical_drive = optical_drive_match.group(1) if optical_drive_match else ''

    return {
        'Имя АРМ': arm_name,
        'Дата отчета': update_date,
        'Пользователь': user,
        'IP-адрес АРМ': ip_address,
        'ОС': os_name,
        'Версия ОС': os_version,
        'Дата установки ОС': os_install_date,
        'Процессор': cpu,
        'Производитель МП': system_manufacturer,
        'Модель МП': model,
        'Всего ОЗУ': total_memory,
        'Занято разъемов ОЗУ': installed_ram_slots,
        'Свободно разъемов ОЗУ': free_ram_slots,
        'Установлено HDD': hdd_count,
        'Объем HDD': hdd_size_gb,
        'CD\DVD': optical_drive,
        'Ссылка': file_path
    }


def main():
    files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.txt')]

    data_list = []
    for file in files:
        data = extract_info(file)
        data_list.append(data)

    df = pd.DataFrame(data_list)

    output_file = 'output.xlsx'
    if os.path.exists(output_file):
        current_time = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
        output_file = f'output_{current_time}.xlsx'

    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        start_row = 1

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        file_path = row[-1]
        row = row[:-1]
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell = ws.cell(row=r_idx, column=len(row) + 1, value=os.path.basename(file_path))
        cell.hyperlink = os.path.abspath(file_path)

    wb.save(output_file)


if __name__ == "__main__":
    main()
